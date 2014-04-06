'''
Created on Jan 3, 2014

@author: jurek
'''
import argparse
import glob
from shutil import copyfile
from openpyxl import Workbook
from openpyxl import load_workbook
from hra_core.io_utils import get_first_lines
from hra_core.io_utils import get_filename
from hra_core.misc import extract_number

# example of eclipse parameters:
################################
#-source_xls '/home/jurek/tmp/klinika0.xlsx'
#-output_xls '/home/jurek/tmp/klinika_out.xlsx'
#-output_sheet_name 'merged'
#-output_file_column_idx 2
#-sources_csv_files_mask '/home/jurek/volumes/doctoral/monitor_do_impedancji_niccomo_wyniki/CO_team/*.res_out',  # @IgnorePep8 one line
#                        '/home/jurek/volumes/doctoral/monitor_do_impedancji_niccomo_wyniki/IVRT_team/*.res_out',  # @IgnorePep8   one line
#                        '/home/jurek/volumes/doctoral/monitor_do_impedancji_niccomo_wyniki/PEP_team/*.res_out',  # @IgnorePep8   one line
#                        '/home/jurek/volumes/doctoral/monitor_do_impedancji_niccomo_wyniki/SV_team/*.res_out',  # @IgnorePep8   one line
#                        '/home/jurek/volumes/doctoral/monitor_do_impedancji_niccomo_wyniki/SVR_team/*.res_out',  # @IgnorePep8   one line
#                        '/home/jurek/volumes/doctoral/monitor_do_impedancji_niccomo_wyniki/LVET_team/*.res_out'  # @IgnorePep8   one line
#-separator ';'
#-index_file_xls '/home/jurek/tmp/indexes.xlsx'
#-index_code_name kodlast7
#-index_sheet_name Arkusz1

#-source_xls '/home/jurek/tmp/healthy_plec_wiek.xlsx'
#-output_xls '/home/jurek/tmp/healthy_plec_wiek_out.xlsx'
#-output_sheet_name 'zdrowi'
#-output_file_column_idx 2
#-sources_csv_files_mask '/home/jurek/volumes/doctoral/monitor_do_impedancji_niccomo_wyniki/CO_healthy/*.res_out',  # @IgnorePep8 one line
#                        '/home/jurek/volumes/doctoral/monitor_do_impedancji_niccomo_wyniki/IVRT_healthy/*.res_out',  # @IgnorePep8   one line
#                        '/home/jurek/volumes/doctoral/monitor_do_impedancji_niccomo_wyniki/PEP_healthy/*.res_out',  # @IgnorePep8   one line
#                        '/home/jurek/volumes/doctoral/monitor_do_impedancji_niccomo_wyniki/SV_healthy/*.res_out',  # @IgnorePep8   one line
#                        '/home/jurek/volumes/doctoral/monitor_do_impedancji_niccomo_wyniki/SVR_healthy/*.res_out'  # @IgnorePep8   one line
#-separator ';'


parser = argparse.ArgumentParser()
parser.add_argument("-source_xls", "--source_xls", help="source xls file")
parser.add_argument("-output_xls", "--output_xls", help="output xls file")
parser.add_argument("-output_sheet_name", "--output_sheet_name",
                    help="output sheet name")
parser.add_argument("-output_file_column_idx", "--output_file_column_idx",
                    help="output file column idx [optional]", default=None,
                    type=int)
parser.add_argument("-index_file_xls", "--index_file_xls",
                    help="xls index file [optional]", default=None)
parser.add_argument("-index_sheet_name", "--index_sheet_name",
                    help="index sheet name [optional]", default=None)
parser.add_argument("-sources_csv_files_mask", "--sources_csv_files_mask",
                    help="source csv files list mask")
parser.add_argument("-separator", "--separator", help="separator")
__args = parser.parse_args()


def merge_xls_by_index_file(output_xls, source_csv_files_mask, separator,
              index_file_xls, index_sheet_name):
    print('***************************************************')
    print('csv source files: ' + source_csv_files_mask)

    index_xls = load_workbook(index_file_xls)
    index_sheet = index_xls.get_sheet_by_name(index_sheet_name)
    index_cells = zip(index_sheet.columns[0], index_sheet.columns[1])

    xls = None
    merged_sheet = None
    for file_nr, _file in enumerate(glob.glob(source_csv_files_mask)):
        #if file_nr == 2:
        #    break
        print('===================================')
        print('Processing file: ' + str(_file))
        if xls == None:
            xls = load_workbook(output_xls)
            merged_sheet = xls.get_sheet_by_name('merged')
            headers_cells = merged_sheet.rows[0]
            code_cells = merged_sheet.columns[2]

        (headers_line, values_line) = get_first_lines(_file, count=2)
        filename = get_filename(_file)
        (group, file_ident) = filename.split('_')
        headers = [str(group + "_" + header.strip())
                        for header in headers_line.split(separator)]
        values = [value.strip() for value in values_line.split(separator)]

        file_number = extract_number(file_ident)
        ident_row = None
        for (file_index_cell, code_index_cell) in index_cells:
            if not ident_row == None:
                break
            if file_number == file_index_cell.value:
                for code_cell in code_cells:
                    if code_cell.value == code_index_cell.value:
                        ident_row = code_cell.row - 1
                        break

        #print("ident_row: " + str(ident_row))
        #add header
        save = False
        if ident_row:
            for header_idx, header in enumerate(headers):

                ident_column = None
                for cell in headers_cells:
                    if header == cell.value:
                        ident_column = cell.column
                        break
                if ident_column == None:
                    ident_column = merged_sheet.get_highest_column()
                    header_cell = merged_sheet.cell(row=0, column=ident_column)
                    header_cell.value = header
                    value_cell = merged_sheet.cell(row=ident_row,
                                               column=ident_column)
                    save = True
                else:
                    value_cell = merged_sheet.cell(
                                coordinate=ident_column + str(ident_row + 1))
                value_cell.value = values[header_idx]
        else:
            print('No item in xls file for ' + str(_file))
            return
        if save:
            print("Saving " + output_xls)
            xls.save(output_xls)
            print(".. done")
            xls = None

    if not xls == None:
        print("Saving " + output_xls)
        xls.save(output_xls)
        print(".. done")


def merge_xls_by_column(output_xls, sheet_name, output_column_idx,
                        source_csv_files_mask, separator):
    print('***************************************************')
    print('csv source files: ' + source_csv_files_mask)

    xls = None
    merged_sheet = None
    for file_nr, _file in enumerate(glob.glob(source_csv_files_mask)):
        #if file_nr == 2:
        #    break
        print('===================================')
        print('Processing file: ' + str(_file))
        if xls == None:
            xls = load_workbook(output_xls)
            merged_sheet = xls.get_sheet_by_name(sheet_name)
            headers_cells = merged_sheet.rows[0]
            code_cells = merged_sheet.columns[output_column_idx]

        (headers_line, values_line) = get_first_lines(_file, count=2)
        filename = get_filename(_file)
        (group, file_ident) = filename.split('_')
        file_ident = file_ident.lstrip('0')
        headers = [str(group + "_" + header.strip())
                        for header in headers_line.split(separator)]
        values = [value.strip() for value in values_line.split(separator)]

        #if file_ident.isdigit():
        #    file_number = int(file_ident)
        #else:
        #    file_number = file_ident.lstrip('0')
        file_number = file_ident

        ident_row = None
        _continue = False
        for code_cell in code_cells:
            value = code_cell.value
            if value == None:
                _continue = True
            elif file_number == value.lstrip('0'):
                ident_row = code_cell.row - 1
                break

        if _continue:
            continue
        #print("ident_row: " + str(ident_row))
        #add header
        save = False
        if ident_row:
            for header_idx, header in enumerate(headers):

                ident_column = None
                for cell in headers_cells:
                    if header == cell.value:
                        ident_column = cell.column
                        break
                if ident_column == None:
                    ident_column = merged_sheet.get_highest_column()
                    header_cell = merged_sheet.cell(row=0, column=ident_column)
                    header_cell.value = header
                    value_cell = merged_sheet.cell(row=ident_row,
                                               column=ident_column)
                    save = True
                else:
                    value_cell = merged_sheet.cell(
                                coordinate=ident_column + str(ident_row + 1))
                value_cell.value = values[header_idx]
        else:
            print('No item in xls file for ' + str(_file))
            continue
        if save:
            print("Saving " + output_xls)
            xls.save(output_xls)
            print(".. done")
            xls = None

    if not xls == None:
        print("Saving " + output_xls)
        xls.save(output_xls)
        print(".. done")


#copy source xml file into output xml file
print("Copy source xml file into output xml file")
copyfile(__args.source_xls, __args.output_xls)
print(".. done")

for source_csv_files_mask in __args.sources_csv_files_mask.split(','):
    if not __args.index_file_xls == None:
        merge_xls_by_index_file(__args.output_xls,
                                #__args.output_sheet_name,
                                #__args.output_file_column_idx,
                                source_csv_files_mask, __args.separator,
                                __args.index_file_xls, __args.index_sheet_name)
    elif not __args.output_file_column_idx == None:
        merge_xls_by_column(__args.output_xls,
                            __args.output_sheet_name,
                            __args.output_file_column_idx,
                            source_csv_files_mask,
                            __args.separator)

print('The End')
